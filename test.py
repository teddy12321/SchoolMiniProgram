import openpyxl
from model.student import Student
from model.classs import Classs
from model.schedule import Schedule
from model.scheduleTemplate import ScheduleTemplate
from data import db
book = openpyxl.load_workbook('sample.xlsx')
sheet = book.active
class1 = Classs(sheet.cell(row=1, column=1).value[2:7]+'A')
s1 = ScheduleTemplate('A',class1,\
                      '高一语文','高一语文','Intro to Writ \n Room 204','ELC1|AP Seminar\nRoom 203','ELC1|AP Seminar\nRoom 203','Pre-Cal\nRoom 208','高一地理','Counseling','高一体育',\
                      '高一物理','TOEFL\nRoom 203','Pre-Phy\nroom 212','Pre-Phy\nRoom 212','高一化学','ELC1|AP Seminar\nRoom 203','班会','大扫除','大扫除',\
                     '高一物理','高一数学','ELC1|AP Seminar\nRoom 203','Pre-Cal\nRoom 208','高一美术/音乐','Intro to Writ \n Room 204','Intro to Writ \n Room 204','高一政治','高一体育',\
                   'Pre-Cal\nRoom 208','高一化学','Intro to Writ \n Room 204','Pre-Phy\nroom 212','Pre-Phy\nroom 212','高一语文','ELC1|AP Seminar\nRoom 203','Self Study','高一历史',\
                      'TOEFL\nRoom 203','Pre-Cal\nRoom 208','Self Study','Intro to Writ \n Room 204','ELC1|AP Seminar\nRoom 203','Pre-Phy\nRoom 212','高一数学','二课堂','二课堂')
class2 = Classs(sheet.cell(row=1, column=1).value[2:7]+'B')
s2 = ScheduleTemplate('B',class2,\
             '高一语文','高一语文','Pre-Phy\nroom 212','Pre-Cal\nRoom 208','TOEFL\nRoom 203','ELC C2','高一地理','Counseling','高一体育',\
            '高一物理','ELC C2','Intro to Writ \n Room 204','Intro to Writ \n Room 204','高一化学','Pre-Cal\nRoom 208','班会','大扫除','大扫除',\
            '高一物理','高一数学','Pre-Cal\nRoom 208','ELC C2','高一美术/音乐','Pre-Phy\nroom 212','Pre-Phy\nroom 212','高一政治','高一体育',\
            'ELC C2','高一化学','Pre-Phy\nroom 212','Intro to Writ \n Room 204','Intro to Writ \n Room 204','高一语文','Pre-Cal\nRoom 208','Self Study','高一历史',\
            'ELC C2','ELC C2','Self Study','Pre-Phy\nroom 212','TOEFL\nRoom 203','Intro to Writ \n Room 204','高一数学','二课堂','二课堂')
class_dic ={"A":class1,"B":class2}
for i in range(3,23):
    name = sheet.cell(row=i, column=2)
    pinYin = sheet.cell(row=i, column=3)
    engName = sheet.cell(row=i, column=4)
    gender = sheet.cell(row=i, column=5)
    group = sheet.cell(row=i, column=7).value
    stuno = sheet.cell(row=i, column=6)
    year = 1
    newstu = Student(name.value, stuno.value, year,None, class_dic[group], gender.value, 0, 0, ' ', ' ', engName.value, pinYin.value)
    db.session.add(newstu)

db.session.add(s1)
db.session.add(s2)
db.session.commit()
