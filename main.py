from flask import  jsonify,request
from pip._vendor import requests
import datetime
from data import app
from data import db
from flask import render_template
import time
from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from model.language import Language
# @app.route('/')
# def root():
#    return 'root'
#
# @app.route('/abc')
# def abc():
#    testDic = {'res':True,'msg':'xxxx','data':[1,2,'aa']}
#    testList = []
#    for i in range(100):
#       testList.append(testDic)
#    return jsonify(testList)
#
from model.absent import Absent
from model.classs import Classs
from model.student import Student
from model.exam import Exam
from model.examprep import ExamPrep
from model.schedule import Schedule
from model.acti import Acti
from model.weather import Weather
from model.delAbsent import DelAbsent
from model.user import User
from model.admin import Admin
from model.checkIn import CheckIn
from model.scheduleTemplate import ScheduleTemplate

STDLATETIME = '07:50:00'


@app.route('/acti')
def recentActi():
   actis = Acti.query.all()
   res = []
   for acti in actis:
      res.append(acti.getDic())
   print(res)
   return jsonify(res)

@app.route('/schedule/<stuno>')
def classSchedule(stuno):
   stu = Student.query.filter(Student.stuNo == stuno).first()
   scheduleId = stu.getSchedule()
   res = {}
   if scheduleId:
      schedule = Schedule.query.filter(Schedule.id == stu.getSchedule()).first()
      res = {"selected": True, "schedule": schedule.construct(), "isStu": True}
   else:
      temp = ScheduleTemplate.query.filter(ScheduleTemplate.classs == stu.classs).first()
      res = {"selected": False, "schedule": temp.construct(), "isStu": True}
   return jsonify(res)

@app.route('/teacherschedule')
def teacherSchedule():
   stus = Student.query.all()
   res = []
   for stu in stus:
      scheduleId = stu.getSchedule()
      if scheduleId:
         schedule = Schedule.query.filter(Schedule.id == stu.getSchedule()).first()
         res.append({"Name": stu.getName(), "schedule": schedule.construct()})
      else:
         res.append({"Name": stu.getName(), "schedule": ''})
   print(res)
   return jsonify(res)

@app.route('/allstu')
def allStu():
   stus = Student.query.all()
   res = []
   for stu in stus:
      res.append({"Name": stu.getName(), "stuNo": stu.getstuNo(), "Class": stu.classs.name})
   return jsonify(res)


@app.route('/exams/<name>')
def exams(name):
   exam1 = Exam.query.filter(Exam.name == name).first()
   print(exam1)
   return jsonify(exam1.getDic())

@app.route('/exams')
def examsall():
   exam1 = Exam.query.all()
   res = []
   for exam in exam1:
      res.append(exam.getDic())
   print(exam1)
   return jsonify(res)

@app.route('/examprep')
def examprep():
   prep1 = ExamPrep.query.filter(ExamPrep.id == 1).first()
   return jsonify(prep1.getDic())


@app.route('/weather')
def weather():
   url = 'http://wthrcdn.etouch.cn/weather_mini?citykey=101030100'
   res = requests.get(url)
   # tdweather = {"date": " ", "weather": " ", "PM2.5": " "}
   return jsonify(res.json())

# @app.route('/login')
# def login():
#    usrname = request.args['usrname']
#    pwd = request.args['pwd']
#    return 0

@app.route('/stuinfo/<stuno>')
def stuinfo(stuno):
   stu1 = Student.query.filter(Student.stuNo==stuno).first()
   print(stu1)
   return jsonify(stu1.getDic())


# @app.route('/applyabsent',methods=['POST'])
# def absent():
#    req = request.json
#    periods = req.get('periods')
#    date = req.get('date')
#    print(date)
#    for period in periods:
#       print(period)
#       absent = Absent(period,date)
#       db.session.add(absent)
#    db.session.commit()
#    return jsonify({'res':True})

@app.route('/checkid', methods = ['POST'])
def checkId():
   req = request.json
   openid = req.get('openid')
   print(openid)
   stuno = ""
   stu = User.query.filter(User.openid==openid).first()
   if(stu != None):
      if(stu.stuNo != None):
         stuno = stu.stuNo
   else:
      newusr = User(openid, None)
      db.session.add(newusr)
      db.session.commit()

   return jsonify({'res':True,
                   'stuno': stuno
                   })

nameList = ['sun', 'mon', 'tue', 'wed', 'thu', 'fri', 'sat']

@app.route('/checkdayschedule', methods = ['POST'])
def checkDaySchedule():
   schList = []
   absList = [False,False,False,False,False,False,False,False,False]

   req = request.json
   stuNo = req.get('stu')
   pos = req.get('day')
   date = req.get('date')
   if(stuNo):
      stu = Student.query.filter(Student.stuNo==stuNo).first()
      type = stu.getSchedule()
      schedule = Schedule.query.filter(Schedule.id == type).first()
      for i in range(1,10):
         schList.append(getattr(schedule,nameList[pos]+str(i),''))
      absents = Absent.query.filter(Absent.Student == stu) \
         .filter(Absent.date == date) \
         .all()
      for absent in absents:
         absList[absent.period-1] = True
      return jsonify({'res': True,
                      'data': schList,
                      'dayabsent': absList
                      })
   else:
      return jsonify({'res': False,
                      })

@app.route('/postid', methods = ['POST'])
def postId():
   req = request.json
   print(req)
   openid = req.get('openid')
   stuno = req.get('stuno')
   name = req.get('name')
   print(stuno)
   print(name)
   stu = Student.query.filter(Student.stuNo == stuno).filter(Student.name == name).first()
   if(stuno=='000000000' and name=='administrator'):
      newusr = User(openid, stuno)
      db.session.add(newusr)
      db.session.commit()
      return jsonify({'res': 'More'})
   if(stu != None):
      usr = User.query.filter(User.openid == openid).first()
      if(usr != None):
         usr.stuNo = stuno
         db.session.commit()
         return jsonify({'res': True})
      else:
         newusr = User(openid, stuno)
         db.session.add(newusr)
         db.session.commit()
         return jsonify({'res': True})
   return jsonify({'res': False})

@app.route('/postapplyabsent', methods = ['POST'])
def postApplyAbsent():
   req = request.json
   print(req)
   stuno = req.get('stuno')
   date = req.get('date')
   classapply = req.get('classapply')
   className = req.get('className')
   print(className)
   stu = Student.query.filter(Student.stuNo == stuno).first()
   for i,period in enumerate(classapply):
      if period:
         abs = Absent.query.filter(Absent.Student == stu) \
            .filter(Absent.date == date) \
            .filter(Absent.period == i + 1).first()
         if not  abs:
            absent = Absent(i+1, date, stu, className[i])
            db.session.add(absent)
            db.session.commit()
      else:
         abs = Absent.query.filter(Absent.Student == stu)\
            .filter(Absent.date == date) \
            .filter(Absent.period == i+1).first()
         if abs:
            db.session.delete(abs)
            db.session.commit()

   return jsonify({'res': True})

@app.route('/admincheck', methods = ['POST'])
def adminCheck():
   req = request.json
   print(req)
   openid = req.get('openid')
   pwd = req.get('pwd')
   if (pwd=='AP2021'):
      newAdmin = Admin(openid)
      db.session.add(newAdmin)
      db.session.commit()
      return jsonify({'res': True})
   return jsonify({'res': False})


@app.route('/findadmin',methods = ['POST'])
def findAdmin():
   req = request.json
   print(req)
   openid = req.get('openid')
   admin = Admin.query.filter(Admin.openid == openid).first()
   if (admin != None):
      return jsonify({'res': True})
   return jsonify({'res': False})


@app.route('/checkin', methods = ['POST'])
def checkIn():
   req = request.json
   print(req)
   stuno = req.get('stuno')
   stu = Student.query.filter(Student.stuNo == stuno).first()
   now = datetime.datetime.now()
   date = now.strftime("%Y-%m-%d")
   time = now.strftime("%H:%M:%S")
   newCheckin = CheckIn(date,time, stu)
   db.session.add(newCheckin)
   db.session.commit()
   return jsonify({'res': True})

@app.route('/checkcheckin', methods = ['POST'])
def checkCheckIn():
   isLate = False
   isChecked = False
   req = request.json
   print(req)
   stuno = req.get('stuno')
   stu = Student.query.filter(Student.stuNo == stuno).first()
   now = datetime.datetime.now()
   date = now.strftime("%Y-%m-%d")
   time = now.strftime("%H:%M:%S")
   check = CheckIn.query.filter(CheckIn.stuid == stu.id).filter(CheckIn.date == date).first()
   if check != None:
      isChecked = True
      if check.time > STDLATETIME:
         isLate = True
   else:
      if time > STDLATETIME:
         isLate = True
   return jsonify({'isChecked': isChecked,
                   'isLate': isLate
                   })

@app.route('/admincheckin')
def adminCheckIn():
   stus = Student.query.all()
   now = datetime.datetime.now()
   date = now.strftime("%Y-%m-%d")
   time = now.strftime("%H:%M:%S")
   latestus = []
   uncheckedstus = []

   for stu in stus:
      isChecked = False
      isLate = False
      check = CheckIn.query.filter(CheckIn.stuid == stu.id).filter(CheckIn.date == date).first()
      if check != None:
         isChecked = True
         if check.time > STDLATETIME:
            isLate = True
      else:
         if time > STDLATETIME:
            isLate = True
      if(isChecked == False):
         uncheckedstus.append({"class": stu.classs.name, "name": stu.getName()})
      elif (isLate == True):
         latestus.append({"class": stu.classs.name, "name": stu.getName()})
   res = {"Late": latestus, "Unchecked" : uncheckedstus}
   return jsonify(res)




@app.route('/scheduletemplate/<stuno>')
def scheduletemplate(stuno):
   stu1 = Student.query.filter(Student.stuNo==stuno).first()
   temps = ScheduleTemplate.query.filter(ScheduleTemplate.classs ==stu1.classs).all()
   res = []
   for temp in temps:
      print(temp.__dict__)
      tmp = temp.__dict__
      tmp.pop('_sa_instance_state')
      res.append(tmp)
   return jsonify(res)

@app.route('/postschedule', methods = ['POST'])
def postschedule():
   req = request.json
   print(req)
   stuno = req.get('stuno')
   schedule = req.get('schedule')
   stu1 = Student.query.filter(Student.stuNo == stuno).first()
   newsch = Schedule(*schedule)
   stu1.schedule = newsch
   db.session.add(newsch)
   db.session.commit()
   return jsonify({'res': True
                   })


@app.route('/getdayabs', methods = ['POST'])
def getDayAbs():
   req = request.json
   date = req.get('date')
   abss = Absent.query.filter(Absent.date == date).all()
   res = []
   for abs in abss:
      res.append({'Name': abs.Student.name, 'ClassName': abs.className, 'id': abs.id, 'Period': abs.period, 'isCompleted': abs.isCompleted})
   return jsonify(res)


@app.route('/postdayabs', methods = ['POST'])
def postDayAbs():
   req = request.json
   id = req.get('id')
   print(id)
   isCompleted = req.get('isCompleted')
   abs = Absent.query.filter(Absent.id == id).first()
   abs.isCompleted = isCompleted
   db.session.commit()
   return jsonify({'res': True
                   })


@app.route('/authorizeall' , methods = ['POST'])
def authorizeAll():
   req = request.json
   date = req.get('date')
   print(date)
   abss = Absent.query.filter(Absent.date == date).all()
   for abs in abss:
      abs.isCompleted = True
   db.session.commit()
   return jsonify({'res': True
                   })


@app.route('/initlanguage' , methods = ['POST'])
def initLanguage():
   req = request.json
   openid = req.get('openid')
   lan = Language.query.filter(Language.openid == openid).first()
   if(lan == None):
      newusr = Language("English", openid)
      db.session.add(newusr)
      db.session.commit()
      return jsonify({'language': newusr.getName()
                      })
   else:
      name = lan.getName()
      return jsonify({'language': name
                   })

@app.route('/changelanguage' , methods = ['POST'])
def changeLanguage():
   req = request.json
   openid = req.get('openid')
   name = req.get('name')
   lan = Language.query.filter(Language.openid == openid).first()
   if(name==lan.name):
      return jsonify({'res': False
                      })
   lan.name = name
   db.session.commit()
   return jsonify({'language': name
                   })


@app.route('/export')
def export():
    return render_template('export.html')

@app.route('/exportdata', methods = ['POST'])
def exportdata():
    start = request.form['start']
    return export_excel(start, '')


def export_excel(start,end):
   """excel ????????????"""
   wb = Workbook()
   sheet = wb.create_sheet(start)
   # sheet = wb.active
   ####################################
   sheet.cell(row=1, column=1).value = "??????"
   sheet.cell(row=1, column=2).value = "??????"
   sheet.cell(row=1, column=3).value = "????????????"
   stus = Student.query.all()
   for i,stu in enumerate(stus):
      sheet.cell(row=i+2, column=1).value = stu.stuNo
      sheet.cell(row=i + 2, column=2).value = stu.getName()
      checkin = CheckIn.query.filter(CheckIn.stuid == stu.id, CheckIn.date == start).first()
      if checkin:
         sheet.cell(row=i + 2, column=3).value = checkin.time
      else:
         sheet.cell(row=i + 2, column=3).value = "?????????"
   ####################################




   # ?????????????????????
   output = BytesIO()
   # ????????????
   wb.save(output)
   output.seek(0)
   filename = "%s.xlsx" % str(int(time.time()))
   fv = send_file(output, as_attachment=True, attachment_filename=filename, conditional=True)
   fv.headers['Content-Disposition'] += "; filename*=utf-8''{}".format(filename)
   fv.headers["Cache-Control"] = "no_store"
   fv.headers["max-age"] = 1


   return fv

if __name__ == '__main__':
   #app.run(host='0.0.0.0',port=443, ssl_context=('top.pem', 'top.key'))
   app.run(host='0.0.0.0',port=80)